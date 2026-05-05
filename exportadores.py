from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Table, TableStyle

from utils import formatar_moeda_br, formatar_percentual_br, sanitizar_dataframe_saida

PASTA_SAIDA = Path("saida")
PASTA_SAIDA.mkdir(exist_ok=True)


def obter_logo_path():
    pasta_base = Path(__file__).resolve().parent

    caminhos_possiveis = [
        pasta_base / "assets" / "logo_empresa.png",
        pasta_base / "assets" / "logo_empresa.jpg",
        pasta_base / "assets" / "logo_empresa.jpeg",
        pasta_base / "logo_empresa.png",
        pasta_base / "logo_empresa.jpg",
        pasta_base / "logo_empresa.jpeg",
    ]

    for caminho in caminhos_possiveis:
        if caminho.exists():
            return caminho

    return None


def salvar_linhas_nao_processadas(falhas, nome_arquivo="linhas_nao_processadas.txt"):
    caminho = PASTA_SAIDA / nome_arquivo
    with open(caminho, "w", encoding="utf-8") as f:
        for linha in falhas:
            f.write(str(linha) + "\n")
    return caminho


def normalizar_cor(valor):
    """
    Normaliza a cor para evitar estouro visual na coluna COR do PDF.
    A ideia é manter apenas a cor base, sem nomes comerciais longos.

    Exemplos:
    - BRANCO BANCHISA  -> BRANCO
    - CINZA SILVERSTONE -> CINZA
    - PRETO VULCANO -> PRETO
    """
    if valor is None:
        return ""

    texto = str(valor).upper().strip()

    if not texto or texto in {"NAN", "NONE", "-"}:
        return ""

    # Remove resíduos que podem aparecer quando o PDF cola a coluna COR com FIPE.
    texto = texto.replace("R$", " ")
    texto = texto.replace("$", " ")
    texto = " ".join(texto.split())

    mapa = {
        "BRANCO BANCHISA": "BRANCO",
        "BRANCO BANQUISE": "BRANCO",
        "BRANCO ATLAS": "BRANCO",
        "BRANCO POLAR": "BRANCO",
        "BRANCO GLACIER": "BRANCO",
        "BRANCO CRISTAL": "BRANCO",
        "BRANCA": "BRANCO",

        "PRETO VULCANO": "PRETO",
        "PRETO ONIX": "PRETO",
        "PRETO CARBON": "PRETO",
        "PRETO NACRE": "PRETO",
        "PRETO OURO": "PRETO",
        "PRETA": "PRETO",

        "CINZA SILVERSTONE": "CINZA",
        "CINZA SILK": "CINZA",
        "CINZA GRANITE": "CINZA",
        "CINZA PLATINUM": "CINZA",
        "CINZA ARTENSE": "CINZA",

        "PRATA BRISK": "PRATA",
        "PRATA BARI": "PRATA",
        "PRATA SAND": "PRATA",
        "PRATA ETOILE": "PRATA",
        "PRATA BILLET": "PRATA",

        "AZUL JAZZ": "AZUL",
        "AZUL BOREAL": "AZUL",
        "AZUL SAPPHIRE": "AZUL",

        "VERMELHO CHILI": "VERMELHO",
        "VERMELHA": "VERMELHO",
        "VERDE SAFARI": "VERDE",
        "AMARELA": "AMARELO",
    }

    for chave, cor_base in mapa.items():
        if texto.startswith(chave):
            return cor_base

    # Fallback: mantém só a primeira palavra, que normalmente é a cor base.
    return texto.split()[0] if texto.split() else ""


def estilizar_planilha(ws):
    fill_header = PatternFill(fill_type="solid", fgColor="D9EAD3")
    font_header = Font(bold=True, color="000000")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    larguras_por_coluna = {
        "PLACA": 14,
        "MODELO": 42,
        "FAB": 8,
        "MOD": 8,
        "KM": 12,
        "COR": 14,
        "FIPE": 16,
        "GANHO IPVA": 16,
        "UF": 8,
        "CIDADE": 32,
        "PREÇO FINAL": 15,
        "DIST FIPE FINAL": 16,
        "MARGEM FINAL": 12,
        "LAUDO CAUTELAR": 16,
        "LINK LAUDO": 18,
    }

    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    for nome_coluna, largura in larguras_por_coluna.items():
        col_idx = header.get(nome_coluna)
        if col_idx:
            letra = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letra].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    fills_colunas = {
        "PREÇO FINAL": PatternFill(fill_type="solid", fgColor="C6E0B4"),
        "DIST FIPE FINAL": PatternFill(fill_type="solid", fgColor="BDD7EE"),
        "MARGEM FINAL": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    }

    for nome_coluna, fill in fills_colunas.items():
        col_idx = header.get(nome_coluna)
        if col_idx:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).fill = fill

    colunas_centradas = {
        "PLACA", "FAB", "MOD", "KM", "COR", "FIPE", "GANHO IPVA", "UF", "CIDADE",
        "PREÇO FINAL", "DIST FIPE FINAL", "MARGEM FINAL", "LAUDO CAUTELAR", "LINK LAUDO",
    }

    for nome_coluna in colunas_centradas:
        col_idx = header.get(nome_coluna)
        if col_idx:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )


def aplicar_formatacao_numerica(ws):
    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    formato_moeda = 'R$ #,##0.00'
    formato_percentual = '0.00%'
    formato_texto = '@'

    for nome_coluna in ["FIPE", "GANHO IPVA", "PREÇO FINAL", "DIST FIPE FINAL"]:
        col_idx = header.get(nome_coluna)
        if col_idx:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = formato_moeda

    col_idx = header.get("MARGEM FINAL")
    if col_idx:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = formato_percentual

    col_km = header.get("KM")
    if col_km:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_km).number_format = '#,##0'

    col_placa = header.get("PLACA")
    if col_placa:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_placa).number_format = formato_texto


def salvar_excel(df: pd.DataFrame, nome_arquivo=None):
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"resultado_atualizado_{timestamp}.xlsx"

    df = sanitizar_dataframe_saida(df)

    if "COR" in df.columns:
        df["COR"] = df["COR"].apply(normalizar_cor)

    caminho_saida = PASTA_SAIDA / nome_arquivo

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="ANÁLISE FINAL", index=False)
        ws = writer.sheets["ANÁLISE FINAL"]

        estilizar_planilha(ws)
        aplicar_formatacao_numerica(ws)

        ws.print_title_rows = "1:1"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

    return caminho_saida


def preparar_dados_pdf(df: pd.DataFrame):
    df_pdf = sanitizar_dataframe_saida(df.copy())

    if "KM" in df_pdf.columns:
        df_pdf["KM"] = df_pdf["KM"].apply(
            lambda x: f"{int(float(x)):,}".replace(",", ".") if pd.notna(x) and str(x).strip() != "" else ""
        )

    for c in ["FIPE", "GANHO IPVA", "PREÇO FINAL", "DIST FIPE FINAL"]:
        if c in df_pdf.columns:
            df_pdf[c] = df_pdf[c].apply(formatar_moeda_br)

    if "MARGEM FINAL" in df_pdf.columns:
        df_pdf["MARGEM FINAL"] = df_pdf["MARGEM FINAL"].apply(formatar_percentual_br)

    if "MODELO" in df_pdf.columns:
        df_pdf["MODELO"] = df_pdf["MODELO"].fillna("").astype(str).str.strip().str.slice(0, 42)

    if "COR" in df_pdf.columns:
        df_pdf["COR"] = df_pdf["COR"].apply(normalizar_cor)

    if "CIDADE" in df_pdf.columns:
        df_pdf["CIDADE"] = df_pdf["CIDADE"].fillna("").astype(str).str.strip().str.slice(0, 34)

    return df_pdf


def _estilos_pdf():
    styles = getSampleStyleSheet()

    return {
        "logo_texto": ParagraphStyle(
            "logo_texto",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=12,
            alignment=TA_CENTER,
            spaceAfter=0,
            spaceBefore=0,
        ),
        "header": ParagraphStyle(
            "header",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=6.2,
            leading=7.2,
            alignment=TA_CENTER,
            spaceAfter=0,
            spaceBefore=0,
        ),
        "texto": ParagraphStyle(
            "texto",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=6.2,
            leading=7.2,
            alignment=TA_LEFT,
            spaceAfter=0,
            spaceBefore=0,
        ),
        "centro": ParagraphStyle(
            "centro",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=6.2,
            leading=7.2,
            alignment=TA_CENTER,
            spaceAfter=0,
            spaceBefore=0,
        ),
        # Fonte menor só para a coluna COR.
        # Essa é a trava principal contra invasão da coluna FIPE.
        "cor": ParagraphStyle(
            "cor",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=5.0,
            leading=5.6,
            alignment=TA_CENTER,
            spaceAfter=0,
            spaceBefore=0,
            wordWrap="CJK",
        ),
    }


def _para(texto, estilo):
    valor = "" if texto is None else str(texto)
    valor = (
        valor.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )
    return Paragraph(valor, estilo)


def _celula_logo(estilos):
    caminho_logo = obter_logo_path()

    if caminho_logo:
        logo = Image(str(caminho_logo))
        logo.drawWidth = 34 * mm
        logo.drawHeight = 12 * mm
        logo.hAlign = "CENTER"
        return logo

    return Paragraph("R3R INTERMEDIAÇÕES", estilos["logo_texto"])


def _detectar_origem_pdf(df: pd.DataFrame) -> str:
    """
    Detecta a origem provável da tabela para aplicar larguras específicas.

    Mantém compatibilidade com qualquer parser:
    - Localiza IPVA costuma ter GANHO IPVA.
    - Movida costuma trazer LAUDO CAUTELAR / LINK LAUDO.
    - Unidas fica como layout próprio quando não há marcadores específicos.
    """
    try:
        for chave in ("origem", "modo", "parser", "fonte"):
            valor = str(df.attrs.get(chave, "") or "").lower()
            if "localiza" in valor:
                return "localiza"
            if "movida" in valor:
                return "movida"
            if "unidas" in valor:
                return "unidas"

        colunas = {str(c).upper().strip() for c in df.columns}

        if "LAUDO CAUTELAR" in colunas or "LINK LAUDO" in colunas:
            return "movida"

        if "GANHO IPVA" in colunas:
            return "localiza"

        if "ARQUIVO ORIGEM" in colunas:
            texto_origem = " ".join(df.get("ARQUIVO ORIGEM", pd.Series(dtype=str)).astype(str).head(20)).lower()
            if "localiza" in texto_origem:
                return "localiza"
            if "movida" in texto_origem:
                return "movida"
            if "unidas" in texto_origem:
                return "unidas"

        return "unidas"
    except Exception:
        return "geral"


def _larguras_pdf_por_coluna(colunas, origem=None):
    """Define larguras por nome de coluna e otimiza o layout por origem."""
    origem = str(origem or "geral").lower().strip()

    mapa_base = {
        "ARQUIVO ORIGEM": 24 * mm,
        "PLACA": 15 * mm,
        "MODELO": 45 * mm,
        "FAB": 8 * mm,
        "MOD": 8 * mm,
        "KM": 11 * mm,
        "COR": 18 * mm,
        "FIPE": 21 * mm,
        "GANHO IPVA": 21 * mm,
        "UF": 7 * mm,
        "CIDADE": 30 * mm,
        "PREÇO FINAL": 21 * mm,
        "DIST FIPE FINAL": 20 * mm,
        "MARGEM FINAL": 14 * mm,
        "LAUDO CAUTELAR": 15 * mm,
        "LINK LAUDO": 22 * mm,
    }

    ajustes_por_origem = {
        # Localiza IPVA: precisa preservar cidade composta e compactar colunas financeiras.
        "localiza": {
            "MODELO": 43 * mm,
            "COR": 17 * mm,
            "FIPE": 21 * mm,
            "GANHO IPVA": 20 * mm,
            "CIDADE": 38 * mm,
            "PREÇO FINAL": 19 * mm,
            "DIST FIPE FINAL": 19 * mm,
            "MARGEM FINAL": 13 * mm,
        },
        # Movida: normalmente exige espaço para laudo/link, então o ajuste é menos agressivo.
        "movida": {
            "MODELO": 43 * mm,
            "COR": 17 * mm,
            "CIDADE": 30 * mm,
            "PREÇO FINAL": 21 * mm,
            "DIST FIPE FINAL": 20 * mm,
            "MARGEM FINAL": 14 * mm,
            "LAUDO CAUTELAR": 17 * mm,
            "LINK LAUDO": 24 * mm,
        },
        # Unidas: tabela tende a ser mais compacta, mas cidade ainda recebe folga.
        "unidas": {
            "MODELO": 46 * mm,
            "COR": 17 * mm,
            "CIDADE": 32 * mm,
            "PREÇO FINAL": 20 * mm,
            "DIST FIPE FINAL": 20 * mm,
            "MARGEM FINAL": 14 * mm,
        },
    }

    mapa = dict(mapa_base)
    mapa.update(ajustes_por_origem.get(origem, {}))

    return [mapa.get(coluna, 18 * mm) for coluna in colunas]


def salvar_pdf(df: pd.DataFrame, nome_arquivo=None, origem=None):
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"resultado_atualizado_{timestamp}.pdf"

    caminho_saida = PASTA_SAIDA / nome_arquivo
    df_pdf = preparar_dados_pdf(df)
    estilos = _estilos_pdf()

    largura_pagina, _ = landscape(A4)
    margem_lateral = 4 * mm

    doc = SimpleDocTemplate(
        str(caminho_saida),
        pagesize=landscape(A4),
        leftMargin=margem_lateral,
        rightMargin=margem_lateral,
        topMargin=4 * mm,
        bottomMargin=4 * mm,
    )

    largura_util = largura_pagina - (margem_lateral * 2)

    origem_layout = origem or _detectar_origem_pdf(df_pdf)
    col_widths = _larguras_pdf_por_coluna(list(df_pdf.columns), origem=origem_layout)
    escala = largura_util / sum(col_widths)
    col_widths = [w * escala for w in col_widths]

    elementos = []
    linhas_por_pagina = 30

    cabecalho = [_para(col, estilos["header"]) for col in df_pdf.columns]

    colunas_centradas = {
        "PLACA", "FAB", "MOD", "KM", "FIPE", "GANHO IPVA", "UF",
        "CIDADE", "PREÇO FINAL", "DIST FIPE FINAL", "MARGEM FINAL",
        "LAUDO CAUTELAR", "LINK LAUDO",
    }

    registros = []
    for _, row in df_pdf.iterrows():
        linha = []
        for coluna in df_pdf.columns:
            if coluna == "COR":
                linha.append(_para(row[coluna], estilos["cor"]))
            else:
                estilo = estilos["centro"] if coluna in colunas_centradas else estilos["texto"]
                linha.append(_para(row[coluna], estilo))
        registros.append(linha)

    for inicio in range(0, len(registros), linhas_por_pagina):
        bloco = registros[inicio:inicio + linhas_por_pagina]

        linha_logo = [_celula_logo(estilos)] + [""] * (len(df_pdf.columns) - 1)
        dados_tabela = [linha_logo, cabecalho] + bloco

        tabela = Table(dados_tabela, colWidths=col_widths, repeatRows=2)

        estilo_tabela = [
            ("SPAN", (0, 0), (-1, 0)),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.white),
            ("TOPPADDING", (0, 0), (-1, 0), 1.2),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 2.0),
            ("LINEBELOW", (0, 0), (-1, 0), 0.35, colors.black),

            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#D9D9D9")),
            ("TEXTCOLOR", (0, 1), (-1, 1), colors.black),

            ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
            ("GRID", (0, 1), (-1, -1), 0.40, colors.black),
            ("LINEBELOW", (0, 1), (-1, 1), 0.55, colors.black),

            ("LEFTPADDING", (0, 1), (-1, -1), 1.2),
            ("RIGHTPADDING", (0, 1), (-1, -1), 1.2),
            ("TOPPADDING", (0, 1), (-1, -1), 1.8),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 1.8),

            ("ROWBACKGROUNDS", (0, 2), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
        ]

        # Destaques por nome da coluna, não por posição fixa.
        for nome_coluna, cor_fundo in {
            "PREÇO FINAL": "#C6E0B4",
            "DIST FIPE FINAL": "#BDD7EE",
            "MARGEM FINAL": "#FFF2CC",
        }.items():
            if nome_coluna in df_pdf.columns:
                idx = list(df_pdf.columns).index(nome_coluna)
                estilo_tabela.append(("BACKGROUND", (idx, 2), (idx, -1), colors.HexColor(cor_fundo)))

        tabela.setStyle(TableStyle(estilo_tabela))
        elementos.append(tabela)

        if inicio + linhas_por_pagina < len(registros):
            elementos.append(PageBreak())

    doc.build(elementos)
    return caminho_saida
