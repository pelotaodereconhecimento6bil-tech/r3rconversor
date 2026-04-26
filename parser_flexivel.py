from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Table, TableStyle

from utils import formatar_moeda_br, formatar_percentual_br, sanitizar_dataframe_saida

PASTA_SAIDA = Path("saida")
PASTA_SAIDA.mkdir(exist_ok=True)


def obter_logo_path():
    pasta_base = Path(__file__).resolve().parent

    caminhos_possiveis = [
        pasta_base / "assets" / "logo_empresa.png",
        pasta_base / "assets" / "logo_empresa.jpg",
        pasta_base / "assets" / "logo_empresa.jpeg",
        pasta_base / "logo_empresa.png",
        pasta_base / "logo_empresa.jpg",
        pasta_base / "logo_empresa.jpeg",
    ]

    for caminho in caminhos_possiveis:
        if caminho.exists():
            return caminho

    return None


def salvar_linhas_nao_processadas(falhas, nome_arquivo="linhas_nao_processadas.txt"):
    caminho = PASTA_SAIDA / nome_arquivo
    with open(caminho, "w", encoding="utf-8") as f:
        for linha in falhas:
            f.write(linha + "\n")
    return caminho


def estilizar_planilha(ws):
    fill_header = PatternFill(fill_type="solid", fgColor="D9EAD3")
    font_header = Font(bold=True, color="000000")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    larguras = {
        "A": 14,
        "B": 42,
        "C": 8,
        "D": 8,
        "E": 12,
        "F": 18,
        "G": 16,
        "H": 8,
        "I": 24,
        "J": 16,
        "K": 18,
        "L": 14,
    }

    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Destaques visuais solicitados:
    # PREÇO FINAL em verde, DIST FIPE FINAL em azul e MARGEM FINAL em amarelo.
    fills_colunas = {
        "J": PatternFill(fill_type="solid", fgColor="C6E0B4"),
        "K": PatternFill(fill_type="solid", fgColor="BDD7EE"),
        "L": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    }

    for col, fill in fills_colunas.items():
        for row in range(2, ws.max_row + 1):
            ws[f"{col}{row}"].fill = fill


def aplicar_formatacao_numerica(ws):
    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    formato_moeda = 'R$ #,##0.00'
    formato_percentual = '0.00%'
    formato_texto = '@'

    for nome_coluna in ["FIPE", "PREÇO FINAL", "DIST FIPE FINAL"]:
        col_idx = header.get(nome_coluna)
        if col_idx:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = formato_moeda

    col_idx = header.get("MARGEM FINAL")
    if col_idx:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = formato_percentual

    col_km = header.get("KM")
    if col_km:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_km).number_format = '#,##0'

    col_placa = header.get("PLACA")
    if col_placa:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_placa).number_format = formato_texto


def salvar_excel(df: pd.DataFrame, nome_arquivo=None):
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"resultado_atualizado_{timestamp}.xlsx"

    df = sanitizar_dataframe_saida(df)
    caminho_saida = PASTA_SAIDA / nome_arquivo

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="ANÁLISE FINAL", index=False)
        ws = writer.sheets["ANÁLISE FINAL"]

        estilizar_planilha(ws)
        aplicar_formatacao_numerica(ws)

        ws.print_title_rows = "1:1"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

    return caminho_saida


def preparar_dados_pdf(df: pd.DataFrame):
    df_pdf = sanitizar_dataframe_saida(df.copy())

    df_pdf["KM"] = df_pdf["KM"].apply(
        lambda x: f"{int(x):,}".replace(",", ".") if pd.notna(x) else ""
    )

    for c in ["FIPE", "PREÇO FINAL", "DIST FIPE FINAL"]:
        df_pdf[c] = df_pdf[c].apply(formatar_moeda_br)

    df_pdf["MARGEM FINAL"] = df_pdf["MARGEM FINAL"].apply(formatar_percentual_br)

    df_pdf["MODELO"] = df_pdf["MODELO"].fillna("").astype(str).str.strip().str.slice(0, 42)
    df_pdf["COR"] = df_pdf["COR"].fillna("").astype(str).str.strip().str.slice(0, 22)
    df_pdf["CIDADE"] = df_pdf["CIDADE"].fillna("").astype(str).str.strip().str.slice(0, 24)

    return df_pdf


def _estilos_pdf():
    styles = getSampleStyleSheet()

    return {
        "logo_texto": ParagraphStyle(
            "logo_texto",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=12,
            alignment=TA_CENTER,
            spaceAfter=0,
            spaceBefore=0,
        ),
        "header": ParagraphStyle(
            "header",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=6.6,
            leading=7.8,
            alignment=TA_CENTER,
            spaceAfter=0,
            spaceBefore=0,
        ),
        "texto": ParagraphStyle(
            "texto",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=6.5,
            leading=7.6,
            alignment=TA_LEFT,
            spaceAfter=0,
            spaceBefore=0,
        ),
        "centro": ParagraphStyle(
            "centro",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=6.5,
            leading=7.6,
            alignment=TA_CENTER,
            spaceAfter=0,
            spaceBefore=0,
        ),
    }


def _para(texto, estilo):
    valor = "" if texto is None else str(texto)
    valor = (
        valor.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )
    return Paragraph(valor, estilo)


def _celula_logo(estilos):
    caminho_logo = obter_logo_path()

    if caminho_logo:
        logo = Image(str(caminho_logo))
        logo.drawWidth = 34 * mm
        logo.drawHeight = 12 * mm
        logo.hAlign = "CENTER"
        return logo

    return Paragraph("R3R INTERMEDIAÇÕES", estilos["logo_texto"])


def salvar_pdf(df: pd.DataFrame, nome_arquivo=None):
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"resultado_atualizado_{timestamp}.pdf"

    caminho_saida = PASTA_SAIDA / nome_arquivo
    df_pdf = preparar_dados_pdf(df)
    estilos = _estilos_pdf()

    largura_pagina, _ = landscape(A4)
    margem_lateral = 4 * mm

    doc = SimpleDocTemplate(
        str(caminho_saida),
        pagesize=landscape(A4),
        leftMargin=margem_lateral,
        rightMargin=margem_lateral,
        topMargin=4 * mm,
        bottomMargin=4 * mm,
    )

    largura_util = largura_pagina - (margem_lateral * 2)

    col_widths = [
        15 * mm, 45 * mm, 9 * mm, 9 * mm, 12 * mm, 22 * mm,
        20 * mm, 9 * mm, 25 * mm, 24 * mm, 25 * mm, 16 * mm,
    ]
    escala = largura_util / sum(col_widths)
    col_widths = [w * escala for w in col_widths]

    elementos = []

    linhas_por_pagina = 31

    cabecalho = [_para(col, estilos["header"]) for col in df_pdf.columns]

    colunas_centradas = {
        "PLACA", "FAB", "MOD", "KM", "FIPE", "UF",
        "CIDADE", "PREÇO FINAL", "DIST FIPE FINAL", "MARGEM FINAL"
    }

    registros = []
    for _, row in df_pdf.iterrows():
        linha = []
        for coluna in df_pdf.columns:
            estilo = estilos["centro"] if coluna in colunas_centradas else estilos["texto"]
            linha.append(_para(row[coluna], estilo))
        registros.append(linha)

    for inicio in range(0, len(registros), linhas_por_pagina):
        bloco = registros[inicio:inicio + linhas_por_pagina]

        linha_logo = [_celula_logo(estilos)] + [""] * (len(df_pdf.columns) - 1)
        dados_tabela = [linha_logo, cabecalho] + bloco

        tabela = Table(dados_tabela, colWidths=col_widths, repeatRows=2)

        tabela.setStyle(TableStyle([
            ("SPAN", (0, 0), (-1, 0)),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.white),
            ("TOPPADDING", (0, 0), (-1, 0), 1.2),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 2.0),
            ("LINEBELOW", (0, 0), (-1, 0), 0.35, colors.black),

            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#D9D9D9")),
            ("TEXTCOLOR", (0, 1), (-1, 1), colors.black),

            ("BACKGROUND", (9, 2), (9, -1), colors.HexColor("#C6E0B4")),
            ("BACKGROUND", (10, 2), (10, -1), colors.HexColor("#BDD7EE")),
            ("BACKGROUND", (11, 2), (11, -1), colors.HexColor("#FFF2CC")),

            ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
            ("GRID", (0, 1), (-1, -1), 0.40, colors.black),
            ("LINEBELOW", (0, 1), (-1, 1), 0.55, colors.black),

            ("LEFTPADDING", (0, 1), (-1, -1), 1.4),
            ("RIGHTPADDING", (0, 1), (-1, -1), 1.4),
            ("TOPPADDING", (0, 1), (-1, -1), 2.0),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 2.0),

            ("ROWBACKGROUNDS", (0, 2), (8, -1), [colors.white, colors.HexColor("#F7F7F7")]),
        ]))

        elementos.append(tabela)

        if inicio + linhas_por_pagina < len(registros):
            elementos.append(PageBreak())

    doc.build(elementos)
    return caminho_saida
