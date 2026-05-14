import re
import textwrap
import base64
from datetime import datetime, timedelta
from pathlib import Path
from core.constantes import COLUNAS_FINAIS
from parsers.localiza import montar_dataframe_localiza
from parsers.movida import montar_dataframe_movida
from parsers.unidas import montar_dataframe_unidas
from parsers.utilitarios import montar_dataframe_utilitarios
from parsers.generico import montar_dataframe_generico
from parsers.tabela_final import (
    montar_dataframe_tabela_final,
    montar_dataframe_tabela_pdf_extraida,
)
from core.normalizador import (
    normalizar_linha,
    normalizar_texto_coluna,
    limpar_cidade,
)
import shutil
import io
import zipfile
from xml.sax.saxutils import escape

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Image, Spacer, Paragraph


from utils import (
    calcular_dist_fipe_final,
    calcular_margem_final,
    calcular_preco_final,
    formatar_moeda_br,
    formatar_percentual_br,
    limpar_km,
    limpar_valor_monetario,
)

PASTA_ENTRADA = Path("entrada")
PASTA_UPLOADS = Path("uploads")
PASTA_SAIDA = Path("saida")
PASTA_HISTORICO = Path("historico")

for _pasta in (PASTA_ENTRADA, PASTA_UPLOADS, PASTA_SAIDA, PASTA_HISTORICO):
    _pasta.mkdir(exist_ok=True)

ARQUIVO_PDF = PASTA_ENTRADA / "SÃO BERNARDO DO CAMPO (PEREIRA BARRETO) 22-04.pdf"


MONEY_PATTERN = r"(?:-?R\$\s?(?:-|\d{1,3}(?:\.\d{3})*(?:,\d{2})?)|Sem FIPE)"
PERCENT_PATTERN = r"(?:-?\d{1,2},\d%|Sem FIPE)"

MAPA_COLUNAS_EQUIVALENTES = {
    "placa": "PLACA",
    "modelo": "MODELO",
    "fab": "FAB",
    "ano fab": "FAB",
    "fabricação": "FAB",
    "mod": "MOD",
    "ano mod": "MOD",
    "modelo ano": "MOD",
    "km": "KM",
    "quilometragem": "KM",
    "cor": "COR",
    "preco": "PRECO",
    "preço": "PRECO",
    "preco cliente": "PRECO",
    "preço cliente": "PRECO",
    "valor": "PRECO",
    "valor venda": "PRECO",
    "orcamento": "ORCAMENTO",
    "orçamento": "ORCAMENTO",
    "fipe": "FIPE",
    "dist fipe": "DIST_FIPE",
    "dif fipe": "DIST_FIPE",
    "diferença fipe": "DIST_FIPE",
    "margem": "DIST_FIPE",
    "marg fipe": "MARGEM_FIPE",
    "margem fipe": "MARGEM_FIPE",
    "% margem": "MARGEM_FIPE",
    "margem %": "MARGEM_FIPE",
    "ganho ipva": "GANHO IPVA",
    "ipva": "GANHO IPVA",
    "uf": "UF",
    "cidade": "CIDADE",
    "endereco": "ENDERECO",
    "endereço": "ENDERECO",
    "laudo cautelar": "LAUDO CAUTELAR",
    "link laudo": "LINK LAUDO",
}


UFS_BRASIL = {
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA",
    "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN",
    "RS", "RO", "RR", "SC", "SP", "SE", "TO",
}
UF_REGEX = "(?:" + "|".join(sorted(UFS_BRASIL)) + ")"




def extrair_linhas_pdf(caminho_pdf: Path):
    linhas = []

    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if not texto:
                continue

            for linha in texto.split("\n"):
                linha = linha.strip()
                if not linha:
                    continue

                if "LOCALIZA SEMINOVOS ATACADO" in linha:
                    continue

                if "Oferta disponível apenas" in linha:
                    continue

                if linha.startswith("Placa Modelo Fab Mod KM Cor Preço"):
                    continue

                if re.match(r"^[A-Z]{3}[0-9][A-Z0-9][0-9]{2}\b", linha):
                    linhas.append(linha)

    return linhas


def extrair_linhas_pdf_flexivel(caminho_pdf: Path):
    linhas = []
    cabecalhos = []

    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if not texto:
                continue

            for linha in texto.split("\n"):
                linha = linha.strip()
                if not linha:
                    continue

                linha_norm = normalizar_texto_coluna(linha)

                if "localiza seminovos atacado" in linha_norm:
                    continue

                if "oferta disponivel apenas" in linha_norm:
                    continue

                if "placa" in linha_norm and "modelo" in linha_norm and "cidade" in linha_norm:
                    cabecalhos.append(linha)
                    continue

                if re.match(r"^[A-Z]{3}[0-9][A-Z0-9][0-9]{2}\b", linha):
                    linhas.append(linha)

    return linhas, cabecalhos


def detectar_colunas_presentes(cabecalhos):
    colunas = set()

    for cab in cabecalhos:
        cab_norm = normalizar_texto_coluna(cab)
        for alias, canonica in MAPA_COLUNAS_EQUIVALENTES.items():
            if alias in cab_norm:
                colunas.add(canonica)

    return colunas



def separar_placa_e_resto(linha_bruta: str):
    match = re.match(
        r"^(?P<placa>[A-Z]{3}[0-9][A-Z0-9][0-9]{2})\s+(?P<resto>.+)$",
        linha_bruta.strip(),
    )
    if not match:
        return None, None

    placa = match.group("placa").strip()
    resto = match.group("resto").strip()
    return placa, resto


def parsear_inicio_sem_placa(resto_inicio: str):
    tokens = resto_inicio.split()

    if len(tokens) < 5:
        return None

    idx_fab = None
    for i in range(len(tokens) - 2):
        if re.fullmatch(r"\d{4}", tokens[i]) and re.fullmatch(r"\d{4}", tokens[i + 1]):
            idx_fab = i
            break

    if idx_fab is None:
        return None

    modelo_tokens = tokens[:idx_fab]
    fab = tokens[idx_fab]
    mod = tokens[idx_fab + 1]

    if idx_fab + 2 >= len(tokens):
        return None

    km = tokens[idx_fab + 2]
    cor_tokens = tokens[idx_fab + 3:]

    if not modelo_tokens or not cor_tokens:
        return None

    return {
        "MODELO": " ".join(modelo_tokens).strip(),
        "FAB": int(fab),
        "MOD": int(mod),
        "KM": limpar_km(km),
        "COR": " ".join(cor_tokens).strip(),
    }


def parsear_linha(linha_bruta: str):
    placa, resto_original = separar_placa_e_resto(linha_bruta)
    if not placa:
        return None

    resto = normalizar_linha(resto_original)

    padrao_final = re.compile(
        rf"(?P<preco>{MONEY_PATTERN})\s+"
        rf"(?P<orcamento>{MONEY_PATTERN})\s+"
        rf"(?P<fipe>{MONEY_PATTERN})\s+"
        rf"(?P<dist>{MONEY_PATTERN})\s+"
        rf"(?P<margem>{PERCENT_PATTERN})\s+"
        rf"(?P<uf>{UF_REGEX})\\s+"
        rf"(?P<local>.+)$",
        re.IGNORECASE,
    )

    match_final = padrao_final.search(resto)
    if not match_final:
        return None

    preco_original = match_final.group("preco")
    fipe = match_final.group("fipe")
    uf = match_final.group("uf")
    local = match_final.group("local").strip()

    inicio_sem_placa = resto[: match_final.start()].strip()

    parte_inicial = parsear_inicio_sem_placa(inicio_sem_placa)
    if not parte_inicial:
        return None

    return {
        "PLACA": placa,
        "MODELO": parte_inicial["MODELO"],
        "FAB": parte_inicial["FAB"],
        "MOD": parte_inicial["MOD"],
        "KM": parte_inicial["KM"],
        "COR": parte_inicial["COR"],
        "FIPE": limpar_valor_monetario(fipe),
        "UF": uf,
        "CIDADE": limpar_cidade(local),
        "PREÇO ORIGINAL": limpar_valor_monetario(preco_original),
    }


def parsear_linha_flexivel(linha_bruta: str, colunas_presentes=None):
    if colunas_presentes is None:
        colunas_presentes = set()

    placa, resto_original = separar_placa_e_resto(linha_bruta)
    if not placa:
        return None

    resto = normalizar_linha(resto_original)
    tokens = resto.split()

    uf_idx = None
    for i in range(len(tokens) - 1, -1, -1):
        token_uf = tokens[i].upper()
        if token_uf in UFS_BRASIL:
            uf_idx = i
            break

    if uf_idx is None:
        return None

    parte_esquerda = " ".join(tokens[:uf_idx]).strip()
    uf = tokens[uf_idx]
    parte_direita = " ".join(tokens[uf_idx + 1:]).strip()

    moedas = re.findall(MONEY_PATTERN, parte_esquerda, flags=re.IGNORECASE)
    percentuais = re.findall(PERCENT_PATTERN, parte_esquerda, flags=re.IGNORECASE)

    if len(moedas) < 2:
        return None

    preco_original = limpar_valor_monetario(moedas[0])
    fipe = limpar_valor_monetario(moedas[2]) if len(moedas) >= 3 else None

    texto_base = parte_esquerda

    if percentuais:
        texto_base = re.sub(
            r"(?:-?\d{1,2},\d%|Sem FIPE)\s*$", "", texto_base, flags=re.IGNORECASE
        ).strip()

    remover_moedas = min(len(moedas), 4)
    for _ in range(remover_moedas):
        texto_base = re.sub(
            r"(?:-?R\$\s?(?:-|\d{1,3}(?:\.\d{3})*(?:,\d{2})?)|Sem FIPE)\s*$",
            "",
            texto_base,
            flags=re.IGNORECASE,
        ).strip()

    tokens_inicio = texto_base.split()

    idx_fab = None
    for i in range(len(tokens_inicio) - 2):
        if re.fullmatch(r"\d{4}", tokens_inicio[i]) and re.fullmatch(r"\d{4}", tokens_inicio[i + 1]):
            idx_fab = i
            break

    if idx_fab is None:
        return None

    modelo_tokens = tokens_inicio[:idx_fab]
    if idx_fab + 2 >= len(tokens_inicio):
        return None

    fab = tokens_inicio[idx_fab]
    mod = tokens_inicio[idx_fab + 1]
    km = tokens_inicio[idx_fab + 2]
    cor_tokens = tokens_inicio[idx_fab + 3:]

    if not modelo_tokens:
        return None

    return {
        "PLACA": placa,
        "MODELO": " ".join(modelo_tokens).strip(),
        "FAB": int(fab),
        "MOD": int(mod),
        "KM": limpar_km(km),
        "COR": " ".join(cor_tokens).strip(),
        "FIPE": fipe,
        "UF": uf,
        "CIDADE": limpar_cidade(parte_direita),
        "PREÇO ORIGINAL": preco_original,
    }
def montar_dataframe_rigido(caminho_pdf: Path):
    linhas = extrair_linhas_pdf(caminho_pdf)

    registros = []
    falhas = []

    for linha in linhas:
        registro = parsear_linha(linha)
        if registro:
            registros.append(registro)
        else:
            falhas.append(linha)

    df = pd.DataFrame(registros)
    return df, falhas


def montar_dataframe_flexivel(caminho_pdf: Path):
    linhas, cabecalhos = extrair_linhas_pdf_flexivel(caminho_pdf)
    colunas_presentes = detectar_colunas_presentes(cabecalhos)

    registros = []
    falhas = []

    for linha in linhas:
        registro = parsear_linha_flexivel(linha, colunas_presentes)
        if registro:
            registros.append(registro)
        else:
            falhas.append(linha)

    df = pd.DataFrame(registros)
    return df, falhas, colunas_presentes


def montar_dataframe_inteligente(caminho_pdf: Path):
    """
    Escolhe automaticamente o melhor modo de leitura.

    Ordem:
    1. Unidas
    2. Movida
    3. Localiza
    4. tabela_pdf
    5. tabela_final
    6. flexível
    7. genérico inteligente
    8. rígido
    """

    df_unidas, falhas_unidas = montar_dataframe_unidas(caminho_pdf)
    total_unidas = len(df_unidas) + len(falhas_unidas)
    taxa_unidas = (len(df_unidas) / total_unidas) if total_unidas else 0

    if len(df_unidas) > 0 and taxa_unidas >= 0.70:
        return df_unidas, falhas_unidas, "unidas", {
            "TIPO VENDA",
            "VALOR FIPE",
            "VALOR C DESCONTO",
        }


    df_utilitarios, falhas_utilitarios = montar_dataframe_utilitarios(caminho_pdf)
    total_utilitarios = len(df_utilitarios) + len(falhas_utilitarios)
    taxa_utilitarios = (len(df_utilitarios) / total_utilitarios) if total_utilitarios else 0

    if len(df_utilitarios) > 0 and taxa_utilitarios >= 0.70:
        return df_utilitarios, falhas_utilitarios, "utilitarios", {
            "PLACA",
            "LOJA",
            "CIDADE",
            "MODELO",
            "FIPE",
            "MARGEM",
            "PREÇO",
            "LAUDO",
            "LINK",
        }


    df_movida, falhas_movida = montar_dataframe_movida(caminho_pdf)
    total_movida = len(df_movida) + len(falhas_movida)
    taxa_movida = (len(df_movida) / total_movida) if total_movida else 0

    if len(df_movida) > 0 and taxa_movida >= 0.70:
        return df_movida, falhas_movida, "movida", {
            "FIPE",
            "ANUNCIO_ATACADO",
            "LAUDO CAUTELAR",
            "LINK LAUDO",
        }

    df_localiza, falhas_localiza, modo_localiza, colunas_localiza = montar_dataframe_localiza(caminho_pdf)

    if len(df_localiza) > 0:
        return df_localiza, falhas_localiza, modo_localiza, colunas_localiza

    df_tabela_pdf, falhas_tabela_pdf = montar_dataframe_tabela_pdf_extraida(caminho_pdf)
    total_tabela_pdf = len(df_tabela_pdf) + len(falhas_tabela_pdf)
    taxa_tabela_pdf = (len(df_tabela_pdf) / total_tabela_pdf) if total_tabela_pdf else 0

    if len(df_tabela_pdf) > 0 and taxa_tabela_pdf >= 0.85:
        return df_tabela_pdf, falhas_tabela_pdf, "tabela_pdf", {"PREÇO_FINAL_INFORMADO"}

    df_final, falhas_final = montar_dataframe_tabela_final(caminho_pdf)
    total_final = len(df_final) + len(falhas_final)
    taxa_final = (len(df_final) / total_final) if total_final else 0

    if len(df_final) > 5 and taxa_final >= 0.85:
        return df_final, falhas_final, "tabela_final", {"PREÇO_FINAL_INFORMADO"}

    df_flex, falhas_flex, colunas_presentes = montar_dataframe_flexivel(caminho_pdf)
    total_flex = len(df_flex) + len(falhas_flex)
    taxa_flex = (len(df_flex) / total_flex) if total_flex else 0

    if len(df_flex) > 0 and taxa_flex >= 0.75:
        return df_flex, falhas_flex, "flexivel", colunas_presentes

    # NOVO MOTOR GENÉRICO
    df_generico, falhas_generico = montar_dataframe_generico(caminho_pdf)
    total_generico = len(df_generico) + len(falhas_generico)
    taxa_generico = (len(df_generico) / total_generico) if total_generico else 0

    if len(df_generico) > 0 and taxa_generico >= 0.60:
        return df_generico, falhas_generico, "generico", {
            "PLACA",
            "MODELO",
            "FIPE",
            "PREÇO ORIGINAL",
        }

    df_rig, falhas_rig = montar_dataframe_rigido(caminho_pdf)
    total_rig = len(df_rig) + len(falhas_rig)
    taxa_rig = (len(df_rig) / total_rig) if total_rig else 0

    if len(df_rig) > 0 and taxa_rig >= 0.75:
        return df_rig, falhas_rig, "rigido", set()

    candidatos = [
        (
            len(df_unidas),
            df_unidas,
            falhas_unidas,
            "unidas",
            {"TIPO VENDA", "VALOR FIPE", "VALOR C DESCONTO"},
        ),

            

        (
            len(df_utilitarios),
            df_utilitarios,
            falhas_utilitarios,
            "utilitarios",
            {"PLACA","LOJA","CIDADE","MODELO","FIPE","MARGEM","PREÇO","LAUDO","LINK",},
        ),


        (
            len(df_movida),
            df_movida,
            falhas_movida,
            "movida",
            {"FIPE", "ANUNCIO_ATACADO", "LAUDO CAUTELAR", "LINK LAUDO"},
        ),

        (
            len(df_tabela_pdf),
            df_tabela_pdf,
            falhas_tabela_pdf,
            "tabela_pdf",
            {"PREÇO_FINAL_INFORMADO"},
        ),

        (
            len(df_final),
            df_final,
            falhas_final,
            "tabela_final",
            {"PREÇO_FINAL_INFORMADO"},
        ),

        (
            len(df_flex),
            df_flex,
            falhas_flex,
            "flexivel",
            colunas_presentes,
        ),

        (
            len(df_generico),
            df_generico,
            falhas_generico,
            "generico",
            {"PLACA", "MODELO", "FIPE", "PREÇO ORIGINAL"},
        ),

        (
            len(df_rig),
            df_rig,
            falhas_rig,
            "rigido",
            set(),
        ),
    ]

    candidatos.sort(key=lambda item: item[0], reverse=True)

    (
        _,
        melhor_df,
        melhores_falhas,
        melhor_modo,
        melhores_colunas,
    ) = candidatos[0]

    return (
        melhor_df,
        melhores_falhas,
        melhor_modo,
        melhores_colunas,
    )

def _coluna_tem_valor_real(df: pd.DataFrame, coluna: str) -> bool:
    """Retorna True quando a coluna existe e tem pelo menos um valor real.

    Usado para impedir que colunas opcionais, como GANHO IPVA,
    sejam criadas artificialmente em listas Movida/Unidas.
    """
    if coluna not in df.columns:
        return False

    serie = df[coluna]
    if serie.empty:
        return False

    serie_txt = serie.fillna("").astype(str).str.strip()
    serie_txt = serie_txt.replace({"nan": "", "None": "", "NONE": "", "NaN": ""})

    return bool(serie_txt.ne("").any())


def aplicar_regras(df: pd.DataFrame, percentual: float = 4.0):
    df = df.copy()

    # GANHO IPVA é opcional. Só deve permanecer se realmente veio do arquivo.
    ganho_ipva_presente = _coluna_tem_valor_real(df, "GANHO IPVA")

    # Regra central: para Movida/Unidas, PREÇO ORIGINAL vem do preço base extraído.
    # As margens originais são ignoradas; a saída é sempre recalculada pelo percentual escolhido.
    if "PREÇO ORIGINAL" not in df.columns and "PREÇO_FINAL_INFORMADO" in df.columns:
        df["PREÇO ORIGINAL"] = df["PREÇO_FINAL_INFORMADO"]

    if "PREÇO ORIGINAL" not in df.columns:
        raise ValueError("Não foi encontrada coluna de preço base/original para recalcular a margem.")

    df["PREÇO FINAL"] = df["PREÇO ORIGINAL"].apply(
        lambda x: calcular_preco_final(x, percentual)
    )

    df["DIST FIPE FINAL"] = df.apply(
        lambda row: calcular_dist_fipe_final(row["FIPE"], row["PREÇO FINAL"]),
        axis=1,
    )

    df["MARGEM FINAL"] = df.apply(
        lambda row: calcular_margem_final(row["DIST FIPE FINAL"], row["FIPE"]),
        axis=1,
    )

    for coluna in ["PREÇO ORIGINAL", "PREÇO_FINAL_INFORMADO"]:
        if coluna in df.columns:
            df.drop(columns=[coluna], inplace=True)

    colunas_saida = list(COLUNAS_FINAIS)

    # Se constantes.py ainda tiver GANHO IPVA, remove quando não veio da origem.
    # Se constantes.py não tiver, insere somente quando veio da origem.
    if ganho_ipva_presente:
        if "GANHO IPVA" not in colunas_saida:
            if "FIPE" in colunas_saida:
                colunas_saida.insert(colunas_saida.index("FIPE") + 1, "GANHO IPVA")
            else:
                colunas_saida.append("GANHO IPVA")
    else:
        if "GANHO IPVA" in colunas_saida:
            colunas_saida.remove("GANHO IPVA")
        if "GANHO IPVA" in df.columns:
            df.drop(columns=["GANHO IPVA"], inplace=True)

    origem = ""
    if "ORIGEM" in df.columns and not df["ORIGEM"].dropna().empty:
        origem = str(df["ORIGEM"].dropna().iloc[0]).lower().strip()

    # Movida e Unidas não usam UF/CIDADE na saída final.
    if origem in {"movida", "unidas"}:
        for coluna_remover in ["UF", "CIDADE"]:
            if coluna_remover in colunas_saida:
                colunas_saida.remove(coluna_remover)

    # Extras específicos da Movida.
    if "LAUDO CAUTELAR" in df.columns or "LINK LAUDO" in df.columns:
        for coluna_extra in ["LAUDO CAUTELAR", "LINK LAUDO"]:
            if coluna_extra not in colunas_saida:
                colunas_saida.append(coluna_extra)

    for coluna in colunas_saida:
        if coluna not in df.columns:
            df[coluna] = "" if coluna in {"LAUDO CAUTELAR", "LINK LAUDO"} else None

    df = df[colunas_saida]

    return df

def salvar_linhas_nao_processadas(falhas, nome_arquivo="linhas_nao_processadas.txt"):
    caminho = PASTA_SAIDA / nome_arquivo
    with open(caminho, "w", encoding="utf-8") as f:
        for linha in falhas:
            f.write(linha + "\n")
    return caminho


def estilizar_planilha(ws):
    fill_header = PatternFill(fill_type="solid", fgColor="D9EAD3")
    font_header = Font(bold=True, color="000000")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    larguras = {
        "ARQUIVO ORIGEM": 28,
        "PLACA": 14,
        "MODELO": 42,
        "FAB": 8,
        "MOD": 8,
        "KM": 12,
        "COR": 18,
        "FIPE": 16,
        "GANHO IPVA": 16,
        "UF": 8,
        "CIDADE": 24,
        "PREÇO FINAL": 16,
        "DIST FIPE FINAL": 18,
        "MARGEM FINAL": 14,
        "LAUDO CAUTELAR": 16,
        "LINK LAUDO": 18,
    }

    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    for nome_coluna, largura in larguras.items():
        col_idx = header.get(nome_coluna)
        if col_idx:
            letra = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letra].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    fills_colunas = {
        "PREÇO FINAL": PatternFill(fill_type="solid", fgColor="C6E0B4"),
        "DIST FIPE FINAL": PatternFill(fill_type="solid", fgColor="BDD7EE"),
        "MARGEM FINAL": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    }

    for nome_coluna, fill in fills_colunas.items():
        col_idx = header.get(nome_coluna)
        if col_idx:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).fill = fill

    # FIPE em negrito no Excel, sem alterar cálculos nem outras colunas.
    col_fipe = header.get("FIPE")
    if col_fipe:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_fipe).font = Font(bold=True)

    for nome_coluna in [
        "ARQUIVO ORIGEM", "PLACA", "FAB", "MOD", "KM", "UF", "FIPE", "GANHO IPVA",
        "PREÇO FINAL", "DIST FIPE FINAL", "MARGEM FINAL",
        "LAUDO CAUTELAR", "LINK LAUDO",
    ]:
        col_idx = header.get(nome_coluna)
        if col_idx:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )


def aplicar_formatacao_numerica(ws):
    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    colunas_moeda = ["FIPE", "GANHO IPVA", "PREÇO FINAL", "DIST FIPE FINAL"]
    colunas_percentual = ["MARGEM FINAL"]

    formato_moeda = 'R$ #,##0.00'
    formato_percentual = '0.00%'
    formato_texto = '@'

    for nome_coluna in colunas_moeda:
        col_idx = header.get(nome_coluna)
        if col_idx:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = formato_moeda

    for nome_coluna in colunas_percentual:
        col_idx = header.get(nome_coluna)
        if col_idx:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = formato_percentual

    col_km = header.get("KM")
    if col_km:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_km).number_format = '#,##0'

    col_placa = header.get("PLACA")
    if col_placa:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_placa).number_format = formato_texto


def salvar_excel(df: pd.DataFrame, nome_arquivo=None):
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"resultado_atualizado_{timestamp}.xlsx"

    caminho_saida = PASTA_SAIDA / nome_arquivo

    df_excel = df.copy()

    links_laudo = None
    if "LINK LAUDO" in df_excel.columns:
        links_laudo = df_excel["LINK LAUDO"].copy()
        df_excel["LINK LAUDO"] = df_excel["LINK LAUDO"].apply(
            lambda x: "Abrir laudo" if str(x or "").startswith("http") else ""
        )

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df_excel.to_excel(writer, sheet_name="ANÁLISE FINAL", index=False)
        ws = writer.sheets["ANÁLISE FINAL"]

        estilizar_planilha(ws)
        aplicar_formatacao_numerica(ws)

        if links_laudo is not None and "LINK LAUDO" in df_excel.columns:
            header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
            col_link = header.get("LINK LAUDO")

            if col_link:
                for row_idx, link in enumerate(links_laudo, start=2):
                    link = str(link or "").strip()
                    if link.startswith("http"):
                        cell = ws.cell(row=row_idx, column=col_link)
                        cell.value = "Abrir laudo"
                        cell.hyperlink = link
                        cell.style = "Hyperlink"

        ws.print_title_rows = "1:1"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.print_options.horizontalCentered = True

    return caminho_saida


def normalizar_cor_pdf(valor):
    texto = str(valor or "").upper().strip()
    texto = re.sub(r"\s+", " ", texto)

    valores_invalidos = {
        "", "-", "--", "---", "NAN", "NONE", "NULL", "N/I", "?",
        "SEM COR", "NÃO INFORMADA", "NAO INFORMADA", "NÃO INFORMADO", "NAO INFORMADO",
    }

    if texto in valores_invalidos:
        return "NÃO INFORMADO"

    texto = texto.replace(" - SÓLIDA", "")
    texto = texto.replace(" - SOLIDA", "")
    texto = texto.replace(" - METÁLICA", "")
    texto = texto.replace(" - METALICA", "")
    texto = texto.replace(" BICOLOR", " C/ TETO PRETO")

    COR_ALIAS = {
        "BRANCA": "BRANCO",
        "PRETA": "PRETO",

        "BRANCO BANQUISE": "BRANCO BANCHISA",
        "BRANCO BANQUISA": "BRANCO BANCHISA",
        "BRANCO BANCHIZA": "BRANCO BANCHISA",

        "BRANCO SUMMIT": "BRANCO SUMMIT",
        "PRETO CARBOM": "PRETO CARBON",
        "SILVER STONE": "CINZA SILVERSTONE",

        "CINZA GRANITE COM TETO PRETO": "CINZA GRANITE C/ TETO PRETO",
        "CINZA GRANITE TETO PRETO": "CINZA GRANITE C/ TETO PRETO",
        "CINZA GRANITE C/ TETO PRETO": "CINZA GRANITE C/ TETO PRETO",
    }

    texto = COR_ALIAS.get(texto, texto)

    CORES_BASE = {
        "BRANCO", "BRANCO BANCHISA", "BRANCO GELEIRA", "BRANCO POLAR", "BRANCO SUMMIT",
        "PRETO", "PRETO VULCANO", "PRETO CARBON", "PRETO MITO", "PRETO PERLA NERA",
        "CINZA", "CINZA SILVERSTONE", "CINZA GRANITE", "CINZA GRANITE C/ TETO PRETO", "CINZA ARTENSE",
        "PRATA", "PRATA BARI", "PRATA BILLET", "PRATA BILLET C/ TETO PRETO",
        "AZUL", "AZUL JAZZ",
        "VERDE",
    }

    if texto in CORES_BASE:
        return texto

    if re.search(r"[A-ZÁÉÍÓÚÃÕÇ]", texto):
        return texto

    return "NÃO INFORMADO"


def obter_estilo_cor(texto):
    tamanho = len(str(texto or ""))

    if tamanho <= 10:
        return 5.8, 6.6
    elif tamanho <= 18:
        return 5.2, 6.0
    elif tamanho <= 26:
        return 4.6, 5.2

    return 4.1, 4.8

def preparar_dados_pdf(df: pd.DataFrame):
    df_pdf = df.copy()

    if "KM" in df_pdf.columns:

        def formatar_km_pdf(x):
            texto = str(x or "").strip()

            if not texto or texto == "-":
                return "-"

            try:
                return f"{int(float(texto)):,}".replace(",", ".")
            except Exception:
                return "-"

        df_pdf["KM"] = df_pdf["KM"].apply(formatar_km_pdf)

    for coluna in ["FIPE", "GANHO IPVA", "PREÇO FINAL", "DIST FIPE FINAL"]:
        if coluna in df_pdf.columns:
            df_pdf[coluna] = df_pdf[coluna].apply(formatar_moeda_br)

    if "MARGEM FINAL" in df_pdf.columns:
        df_pdf["MARGEM FINAL"] = df_pdf["MARGEM FINAL"].apply(formatar_percentual_br)

    if "MODELO" in df_pdf.columns:
        df_pdf["MODELO"] = (
            df_pdf["MODELO"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.slice(0, 40)
        )

    if "COR" in df_pdf.columns:
        df_pdf["COR"] = (
            df_pdf["COR"]
            .fillna("")
            .astype(str)
            .str.strip()
            .apply(normalizar_cor_pdf)
        )

    def limpar_cidade_pdf(valor):
        texto = str(valor or "").strip()
        texto = re.sub(r"\s+", " ", texto)

        if not texto:
            return ""

        texto_upper = texto.upper()

        cidades_conhecidas = [
            "SÃO BERNARDO DO CAMPO",
            "SAO BERNARDO DO CAMPO",
            "SÃO JOSÉ DO RIO PRETO",
            "SAO JOSE DO RIO PRETO",
            "SÃO JOSÉ DOS CAMPOS",
            "SAO JOSE DOS CAMPOS",
            "PRESIDENTE PRUDENTE",
            "RIBEIRÃO PRETO",
            "RIBEIRAO PRETO",
            "SANTO ANDRÉ",
            "SANTO ANDRE",
            "SÃO PAULO",
            "SAO PAULO",
            "GUARULHOS",
            "PIRACICABA",
            "SOROCABA",
            "BAURU",
            "OSASCO",
            "CAMPINAS",
            "BARUERI",
        ]

        cidade_corrigida = texto_upper

        for cidade in sorted(cidades_conhecidas, key=len, reverse=True):
            if texto_upper.startswith(cidade):
                cidade_corrigida = cidade
                break

        correcoes = {
            "SAO PAULO": "SÃO PAULO",
            "SAO BERNARDO DO CAMPO": "SÃO BERNARDO DO CAMPO",
            "SAO JOSE DO RIO PRETO": "SÃO JOSÉ DO RIO PRETO",
            "SAO JOSE DOS CAMPOS": "SÃO JOSÉ DOS CAMPOS",
            "RIBEIRAO PRETO": "RIBEIRÃO PRETO",
            "SANTO ANDRE": "SANTO ANDRÉ",
        }

        return correcoes.get(cidade_corrigida, cidade_corrigida)

    if "CIDADE" in df_pdf.columns:
        df_pdf["CIDADE"] = df_pdf["CIDADE"].apply(limpar_cidade_pdf)

    for coluna in ["LAUDO CAUTELAR", "LINK LAUDO"]:
        if coluna in df_pdf.columns:
            df_pdf[coluna] = (
                df_pdf[coluna]
                .fillna("")
                .astype(str)
                .str.strip()
            )

    return df_pdf


def obter_logo_streamlit_path():
    caminhos = [
        Path("assets/land_icon.png"),
        Path(__file__).resolve().parent / "assets" / "land_icon.png",
    ]

    for caminho in caminhos:
        if caminho.exists():
            return caminho

    return None


def obter_logo_pdf_path():
    caminhos = [
        Path("assets/margem_icon.png"),
        Path(__file__).resolve().parent / "assets" / "margem_icon.png",
    ]

    for caminho in caminhos:
        if caminho.exists():
            return caminho

    return None



def _formatar_celula_pdf(valor, coluna):
    valor_txt = str(valor or "").strip()

    if coluna == "LINK LAUDO" and valor_txt.startswith("http"):
        url = escape(valor_txt, {'"': '&quot;'})
        return Paragraph(
            f'<link href="{url}"><font color="blue" size="8"><u>Abrir laudo</u></font></link>'
        )

    return valor_txt


def salvar_pdf(df: pd.DataFrame, nome_arquivo=None):
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"resultado_atualizado_{timestamp}.pdf"

    caminho_saida = PASTA_SAIDA / nome_arquivo

    df_pdf = preparar_dados_pdf(df).copy()

    origem_pdf = str(df.attrs.get("origem", "") or "").lower().strip()

    if "ORIGEM" in df_pdf.columns:
        origem_pdf = (
            df_pdf["ORIGEM"]
            .fillna("")
            .astype(str)
            .str.lower()
            .iloc[0]
        )

    movida = (
        "LAUDO CAUTELAR" in df_pdf.columns
        or "LINK LAUDO" in df_pdf.columns
    )

    if movida:
        colunas_pdf = [
            "PLACA", "MODELO", "FAB", "MOD", "KM", "COR", "FIPE",
            "PREÇO FINAL", "DIST FIPE FINAL", "MARGEM FINAL",
            "LAUDO CAUTELAR", "LINK LAUDO",
        ]

        # GANHO IPVA só entra no PDF se realmente existir no DataFrame.
        if _coluna_tem_valor_real(df_pdf, "GANHO IPVA"):
            colunas_pdf.insert(colunas_pdf.index("PREÇO FINAL"), "GANHO IPVA")

        for coluna in colunas_pdf:
            if coluna not in df_pdf.columns:
                df_pdf[coluna] = ""
        df_pdf = df_pdf[colunas_pdf]
    else:
        if "MODELO" in df_pdf.columns:
            df_pdf["MODELO"] = (
                df_pdf["MODELO"]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.slice(0, 40)
            )

    largura_pagina, _ = landscape(A4)
    margem = 4 * mm

    doc = SimpleDocTemplate(
        str(caminho_saida),
        pagesize=landscape(A4),
        leftMargin=margem,
        rightMargin=margem,
        topMargin=5 * mm,
        bottomMargin=4 * mm,
    )

    largura_util = largura_pagina - (margem * 2)

    # Largura dinâmica por nome da coluna.
    # Isso evita desalinhamento quando UF/CIDADE são removidas na Unidas/Movida.
    larguras_por_coluna = {
        "ARQUIVO ORIGEM": 26 * mm,
        "PLACA": 15 * mm,
        "MODELO": 58 * mm,
        "FAB": 8 * mm,
        "MOD": 8 * mm,
        "KM": 11 * mm,
        "COR": 18 * mm,
        "FIPE": 24 * mm,
        "GANHO IPVA": 24 * mm,
        "UF": 7 * mm,
        "CIDADE": 30 * mm,
        "PREÇO FINAL": 32 * mm,
        "DIST FIPE FINAL": 29 * mm,
        "MARGEM FINAL": 18 * mm,
        "LAUDO CAUTELAR": 15 * mm,
        "LINK LAUDO": 22 * mm,
    }

    if origem_pdf == "utilitarios":

        larguras_por_coluna.update({

        "PREÇO FINAL": 26 * mm,

        "DIST FIPE FINAL": 24 * mm,

        "MARGEM FINAL": 16 * mm,

        "LAUDO CAUTELAR": 28 * mm,

        "LINK LAUDO": 14 * mm,
    })

    col_widths = [
        larguras_por_coluna.get(coluna, 18 * mm)
        for coluna in df_pdf.columns
    ]

    linhas_por_pagina = 28 if movida else 30

    escala = largura_util / sum(col_widths)
    col_widths = [w * escala for w in col_widths]

    elementos = []
    cabecalho = list(df_pdf.columns)

    registros = []
    for _, row in df_pdf.iterrows():
        registros.append([
            _formatar_celula_pdf(row[coluna], coluna)
            for coluna in cabecalho
        ])

    logo_path = obter_logo_pdf_path()

    for inicio in range(0, len(registros), linhas_por_pagina):
        bloco = registros[inicio:inicio + linhas_por_pagina]

        if logo_path:
            logo = Image(str(logo_path))
            logo.drawWidth = 48 * mm
            logo.drawHeight = 22 * mm
            logo.hAlign = "CENTER"
            elementos.append(logo)
        else:
            elementos.append(Paragraph("<b>R3R INTERMEDIAÇÕES</b>"))

        elementos.append(Spacer(1, 4 * mm))

        tabela = Table(
            [cabecalho] + bloco,
            colWidths=col_widths,
            repeatRows=1,
        )

        estilos = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#BFBFBF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5 if movida else 6.8),
            ("LEADING", (0, 0), (-1, -1), 8.8 if movida else 9.3),
            ("LEFTPADDING", (0, 0), (-1, -1), 1.4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1.4),
            ("TOPPADDING", (0, 0), (-1, -1), 2.0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.0),
            ("GRID", (0, 0), (-1, -1), 0.32, colors.HexColor("#7A7A7A")),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.black),
            ("LINEBELOW", (0, 0), (-1, 0), 0.55, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                colors.white,
                colors.HexColor("#FBFBFB"),
            ]),
        ]

        for nome_coluna, cor_fundo in {
            "PREÇO FINAL": "#C6E0B4",
            "DIST FIPE FINAL": "#BDD7EE",
            "MARGEM FINAL": "#FFF2CC",
        }.items():
            if nome_coluna in cabecalho:
                idx_col = cabecalho.index(nome_coluna)
                estilos.append(("BACKGROUND", (idx_col, 1), (idx_col, -1), colors.HexColor(cor_fundo)))

        for nome_coluna in [
            "ARQUIVO ORIGEM", "PLACA", "FAB", "MOD", "KM", "FIPE", "GANHO IPVA", "UF",
            "PREÇO FINAL", "DIST FIPE FINAL", "MARGEM FINAL",
            "LAUDO CAUTELAR", "LINK LAUDO",
        ]:
            if nome_coluna in cabecalho:
                idx_col = cabecalho.index(nome_coluna)
                estilos.append(("ALIGN", (idx_col, 1), (idx_col, -1), "CENTER"))

        # FIPE em negrito no PDF para todos os modelos, usando nome da coluna.
        if "FIPE" in cabecalho:
            idx_fipe = cabecalho.index("FIPE")
            estilos.append(("FONTNAME", (idx_fipe, 1), (idx_fipe, -1), "Helvetica-Bold"))

        if "MODELO" in cabecalho:
            idx_modelo = cabecalho.index("MODELO")
            estilos.append(("ALIGN", (idx_modelo, 1), (idx_modelo, -1), "LEFT"))

        if "COR" in cabecalho:
            idx_cor = cabecalho.index("COR")

            for linha_pdf in range(1, len(bloco) + 1):
                valor_cor = bloco[linha_pdf - 1][idx_cor]
                fonte, leading = obter_estilo_cor(valor_cor)

                estilos.append(("FONTSIZE", (idx_cor, linha_pdf), (idx_cor, linha_pdf), fonte))
                estilos.append(("LEADING", (idx_cor, linha_pdf), (idx_cor, linha_pdf), leading))

            estilos.append(("ALIGN", (idx_cor, 1), (idx_cor, -1), "CENTER"))

        if "CIDADE" in cabecalho:
            idx_cidade = cabecalho.index("CIDADE")
            estilos.append(("FONTSIZE", (idx_cidade, 1), (idx_cidade, -1), 5.6))
            estilos.append(("LEADING", (idx_cidade, 1), (idx_cidade, -1), 6.3))
            estilos.append(("ALIGN", (idx_cidade, 1), (idx_cidade, -1), "CENTER"))

        tabela.setStyle(TableStyle(estilos))
        elementos.append(tabela)

        if inicio + linhas_por_pagina < len(registros):
            elementos.append(PageBreak())

    doc.build(elementos)
    return caminho_saida



MAPA_COLUNAS_EXCEL = {
    "placa": "PLACA",
    "modelo": "MODELO",
    "fab": "FAB",
    "ano fab": "FAB",
    "ano fabricacao": "FAB",
    "ano fabricação": "FAB",
    "mod": "MOD",
    "ano mod": "MOD",
    "ano modelo": "MOD",
    "km": "KM",
    "quilometragem": "KM",
    "cor": "COR",
    "fipe": "FIPE",
    "ganho ipva": "GANHO IPVA",
    "ganho do ipva": "GANHO IPVA",
    "ipva": "GANHO IPVA",
    "uf": "UF",
    "cidade": "CIDADE",
    "preco": "PREÇO ORIGINAL",
    "preço": "PREÇO ORIGINAL",
    "valor": "PREÇO ORIGINAL",
    "valor venda": "PREÇO ORIGINAL",
    "preco original": "PREÇO ORIGINAL",
    "preço original": "PREÇO ORIGINAL",
    "preco final": "PREÇO ORIGINAL",
    "preço final": "PREÇO ORIGINAL",
    "laudo cautelar": "LAUDO CAUTELAR",
    "link laudo": "LINK LAUDO",
}


def nome_seguro(nome: str) -> str:
    nome = str(nome or "").strip()
    if not nome:
        nome = "resultado_margem_atualizada"

    proibidos = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
    for char in proibidos:
        nome = nome.replace(char, "_")

    nome = "_".join(nome.split())
    return nome[:90] or "resultado_margem_atualizada"


def limpar_historico_antigo(dias: int = 10):
    limite = datetime.now() - timedelta(days=dias)

    for pasta_base in (PASTA_HISTORICO, PASTA_UPLOADS):
        if not pasta_base.exists():
            continue

        for item in pasta_base.iterdir():
            try:
                data_mod = datetime.fromtimestamp(item.stat().st_mtime)
                if data_mod < limite:
                    if item.is_dir():
                        shutil.rmtree(item, ignore_errors=True)
                    else:
                        item.unlink(missing_ok=True)
            except Exception:
                pass


def salvar_upload(uploaded_file) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome = nome_seguro(uploaded_file.name)
    caminho = PASTA_UPLOADS / f"{timestamp}_{nome}"

    with open(caminho, "wb") as f:
        f.write(uploaded_file.getbuffer())

    return caminho


def ler_bytes(caminho: Path):
    caminho = Path(caminho)

    if not caminho.exists():
        return None

    with open(caminho, "rb") as f:
        return f.read()


def normalizar_nome_coluna_excel(coluna: str) -> str:
    texto = str(coluna).strip().lower()
    trocas = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    for origem, destino in trocas.items():
        texto = texto.replace(origem, destino)
    texto = re.sub(r"\s+", " ", texto)
    return texto


def padronizar_excel(df: pd.DataFrame) -> pd.DataFrame:
    renomear = {}

    for coluna in df.columns:
        chave = normalizar_nome_coluna_excel(coluna)
        if chave in MAPA_COLUNAS_EXCEL:
            renomear[coluna] = MAPA_COLUNAS_EXCEL[chave]
        else:
            coluna_upper = str(coluna).strip().upper()
            if coluna_upper in COLUNAS_FINAIS or coluna_upper in {"PREÇO ORIGINAL", "PREÇO_FINAL_INFORMADO"}:
                renomear[coluna] = coluna_upper

    df = df.rename(columns=renomear).copy()

    if "PLACA" not in df.columns:
        raise ValueError("A planilha precisa ter uma coluna de placa.")

    if "PREÇO ORIGINAL" not in df.columns and "PREÇO_FINAL_INFORMADO" not in df.columns:
        raise ValueError("A planilha precisa ter coluna de preço, valor, preço original ou preço final.")

    for coluna in ["FIPE", "GANHO IPVA", "PREÇO ORIGINAL", "PREÇO_FINAL_INFORMADO"]:
        if coluna in df.columns:
            df[coluna] = df[coluna].apply(limpar_valor_monetario)

    if "KM" in df.columns:
        df["KM"] = df["KM"].apply(limpar_km)

    for coluna in ["MODELO", "COR", "UF", "CIDADE"]:
        if coluna not in df.columns:
            df[coluna] = ""

    for coluna in ["FAB", "MOD", "KM", "FIPE"]:
        if coluna not in df.columns:
            df[coluna] = None

    colunas_base = [
        "PLACA", "MODELO", "FAB", "MOD", "KM", "COR",
        "FIPE", "UF", "CIDADE"
    ]

    if "GANHO IPVA" in df.columns:
        colunas_base.append("GANHO IPVA")

    if "PREÇO ORIGINAL" in df.columns:
        colunas_base.append("PREÇO ORIGINAL")
    if "PREÇO_FINAL_INFORMADO" in df.columns:
        colunas_base.append("PREÇO_FINAL_INFORMADO")

    for coluna_extra in ["LAUDO CAUTELAR", "LINK LAUDO"]:
        if coluna_extra in df.columns:
            colunas_base.append(coluna_extra)

    return df[colunas_base].copy()


def processar_arquivo(caminho: Path, percentual: float):
    ext = caminho.suffix.lower()

    if ext == ".pdf":
        df, falhas, modo_usado, colunas_detectadas = montar_dataframe_inteligente(caminho)
    elif ext in {".xlsx", ".xls"}:
        df = pd.read_excel(caminho)
        df = padronizar_excel(df)
        falhas = []
        modo_usado = "excel"
        colunas_detectadas = set(df.columns)
    else:
        raise ValueError("Formato não suportado. Envie PDF, XLSX ou XLS.")

    if df.empty:
        return pd.DataFrame(), {
            "modo": modo_usado,
            "falhas": falhas,
            "colunas_detectadas": sorted(list(colunas_detectadas)) if colunas_detectadas else [],
        }

    df_final = aplicar_regras(df, percentual=percentual)
    df_final.attrs["origem"] = modo_usado

    return df_final, {
        "modo": modo_usado,
        "falhas": falhas,
        "colunas_detectadas": sorted(list(colunas_detectadas)) if colunas_detectadas else [],
    }


def salvar_historico(caminho_excel: Path, caminho_pdf: Path, nome_base: str):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pasta = PASTA_HISTORICO / f"{timestamp}_{nome_seguro(nome_base)}"
    pasta.mkdir(parents=True, exist_ok=True)

    destino_excel = pasta / f"{nome_base}.xlsx"
    destino_pdf = pasta / f"{nome_base}.pdf"

    shutil.copy2(caminho_excel, destino_excel)
    shutil.copy2(caminho_pdf, destino_pdf)

    return destino_excel, destino_pdf


def listar_historico():
    itens = []

    if not PASTA_HISTORICO.exists():
        return itens

    for pasta in sorted(PASTA_HISTORICO.iterdir(), reverse=True):
        if not pasta.is_dir():
            continue

        excel = next((p for p in pasta.glob("*.xlsx")), None)
        pdf = next((p for p in pasta.glob("*.pdf")), None)

        if excel or pdf:
            itens.append({
                "pasta": pasta,
                "excel": excel,
                "pdf": pdf,
                "data": datetime.fromtimestamp(pasta.stat().st_mtime),
            })

    return itens


def excluir_itens_historico(pastas):
    excluidos = 0

    for pasta in pastas:
        try:
            pasta = Path(pasta)
            if pasta.exists() and pasta.is_dir() and pasta.parent.resolve() == PASTA_HISTORICO.resolve():
                shutil.rmtree(pasta, ignore_errors=True)
                excluidos += 1
        except Exception:
            pass

    return excluidos



def listar_uploads_historico():
    itens = []

    if not PASTA_UPLOADS.exists():
        return itens

    for arquivo in sorted(PASTA_UPLOADS.iterdir(), reverse=True):
        if not arquivo.is_file():
            continue

        if arquivo.suffix.lower() not in {".pdf", ".xlsx", ".xls"}:
            continue

        itens.append({
            "arquivo": arquivo,
            "data": datetime.fromtimestamp(arquivo.stat().st_mtime),
            "tamanho": arquivo.stat().st_size,
        })

    return itens


def formatar_tamanho_arquivo(tamanho_bytes: int) -> str:
    tamanho = float(tamanho_bytes)

    for unidade in ["B", "KB", "MB", "GB"]:
        if tamanho < 1024:
            return f"{tamanho:.1f} {unidade}".replace(".", ",")
        tamanho /= 1024

    return f"{tamanho:.1f} TB".replace(".", ",")


def excluir_uploads_historico(arquivos):
    excluidos = 0

    for arquivo in arquivos:
        try:
            arquivo = Path(arquivo)
            if arquivo.exists() and arquivo.is_file() and arquivo.parent.resolve() == PASTA_UPLOADS.resolve():
                arquivo.unlink(missing_ok=True)
                excluidos += 1
        except Exception:
            pass

    return excluidos


def _paginador_itens(itens, chave_base: str, itens_por_pagina: int = 10):
    total_itens = len(itens)
    total_paginas = max(1, (total_itens + itens_por_pagina - 1) // itens_por_pagina)

    pagina_atual = st.session_state.get(chave_base, 1)
    pagina_atual = max(1, min(pagina_atual, total_paginas))

    inicio = (pagina_atual - 1) * itens_por_pagina
    fim = inicio + itens_por_pagina

    return itens[inicio:fim], pagina_atual, total_paginas


def _renderizar_paginacao(chave_base, pagina_atual, total_paginas):

    if total_paginas <= 1:
        return

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    largura = min(total_paginas, 10)

    espacos = [0.42]

    for _ in range(largura):
        espacos.append(0.045)

    espacos.append(0.42)

    cols = st.columns(espacos)

    for pagina in range(1, total_paginas + 1):

        col = cols[pagina]

        with col:

            if st.button(
                str(pagina),
                key=f"{chave_base}_{pagina}",
                type="primary" if pagina == pagina_atual else "secondary",
                use_container_width=True,
            ):
                st.session_state[chave_base] = pagina
                st.rerun()


def renderizar_historico_uploads():
    with st.expander("📁 Histórico de arquivos enviados — últimos 10 dias", expanded=False):
        itens = listar_uploads_historico()

        if not itens:
            st.caption("Nenhum arquivo enviado no histórico.")
            return

        itens_pagina, pagina_atual, total_paginas = _paginador_itens(
            itens,
            chave_base="pagina_uploads",
            itens_por_pagina=10,
        )

        st.caption(f"{len(itens)} arquivo(s) enviado(s) encontrados.")

        col_a, col_b, col_c = st.columns([0.18, 0.18, 0.64])

        with col_a:
            if st.button(
                "Selecionar página",
                use_container_width=True,
                key="btn_select_uploads_pagina",
            ):
                for item in itens_pagina:
                    arquivo = item["arquivo"]
                    st.session_state[f"selecionar_upload_{arquivo.name}"] = True
                st.rerun()

        with col_b:
            if st.button(
                "Limpar página",
                use_container_width=True,
                key="btn_clear_uploads_pagina",
            ):
                for item in itens_pagina:
                    arquivo = item["arquivo"]
                    st.session_state[f"selecionar_upload_{arquivo.name}"] = False
                st.rerun()

        selecionados = []

        for item in itens_pagina:
            arquivo = item["arquivo"]
            data_txt = item["data"].strftime("%d/%m/%Y %H:%M")
            tamanho_txt = formatar_tamanho_arquivo(item["tamanho"])
            chave = f"selecionar_upload_{arquivo.name}"

            with st.container(border=True):
                col_sel, col_info, col_download = st.columns([0.05, 0.70, 0.25])

                with col_sel:
                    marcado = st.checkbox(
                        "Selecionar arquivo enviado",
                        key=chave,
                        label_visibility="collapsed",
                    )

                with col_info:
                    st.markdown(
                        f"""
                        <div style="padding-top:6px; line-height:1.25;">
                            <div style="font-weight:650; font-size:15px;">
                                {arquivo.name}
                            </div>
                            <div style="opacity:0.68; font-size:12px; padding-top:3px;">
                                {data_txt} · {tamanho_txt}
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                with col_download:
                    mime = (
                        "application/pdf"
                        if arquivo.suffix.lower() == ".pdf"
                        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    arquivo_bytes = ler_bytes(arquivo)

                    if arquivo_bytes:
                        st.download_button(
                            "⬇ Baixar",
                            data=arquivo_bytes,
                            file_name=arquivo.name,
                            mime=mime,
                            key=f"baixar_upload_{arquivo.name}",
                            use_container_width=True,
                        )
                    else:
                        st.warning("Arquivo não encontrado.")

                if marcado:
                    selecionados.append(arquivo)

        _renderizar_paginacao("pagina_uploads", pagina_atual, total_paginas)

        st.caption(f"Selecionados nesta página: {len(selecionados)}")

        if st.button(
            "🗑️ Excluir enviados selecionados",
            type="secondary",
            use_container_width=True,
            disabled=len(selecionados) == 0,
            key="excluir_uploads_selecionados",
        ):
            qtd = excluir_uploads_historico(selecionados)

            for arquivo in selecionados:
                st.session_state.pop(f"selecionar_upload_{Path(arquivo).name}", None)

            st.success(f"{qtd} arquivo(s) enviado(s) excluído(s).")
            st.rerun()


def _renderizar_previa_historico_excel(caminho_excel: Path, chave: str):
    """Mostra uma prévia rápida do Excel já convertido no histórico."""
    if not caminho_excel or not Path(caminho_excel).exists():
        st.caption("Prévia indisponível: Excel não encontrado.")
        return

    try:
        df_prev = pd.read_excel(caminho_excel)
    except Exception as e:
        st.warning(f"Não foi possível abrir a prévia deste arquivo: {e}")
        return

    if df_prev.empty:
        st.caption("Arquivo sem registros para pré-visualização.")
        return

    total = len(df_prev)
    st.caption(f"Pré-visualização: {min(total, 50)} de {total} registro(s).")
    st.dataframe(df_prev.head(50), use_container_width=True, height=260)


def renderizar_historico():
    with st.expander("📊 Histórico de arquivos convertidos — últimos 10 dias", expanded=False):
        itens = listar_historico()

        if not itens:
            st.caption("Nenhum arquivo no histórico.")
            return

        itens_pagina, pagina_atual, total_paginas = _paginador_itens(
            itens,
            chave_base="pagina_convertidos",
            itens_por_pagina=10,
        )

        st.caption(f"{len(itens)} conversão(ões) encontradas.")

        col_a, col_b, col_c = st.columns([0.18, 0.18, 0.64])

        with col_a:
            if st.button(
                "Selecionar página",
                use_container_width=True,
                key="btn_select_convertidos_pagina",
            ):
                for item in itens_pagina:
                    pasta = item["pasta"]
                    st.session_state[f"selecionar_historico_{pasta.name}"] = True
                st.rerun()

        with col_b:
            if st.button(
                "Limpar página",
                use_container_width=True,
                key="btn_clear_convertidos_pagina",
            ):
                for item in itens_pagina:
                    pasta = item["pasta"]
                    st.session_state[f"selecionar_historico_{pasta.name}"] = False
                st.rerun()

        selecionados = []

        for item in itens_pagina:
            pasta = item["pasta"]
            data_txt = item["data"].strftime("%d/%m/%Y %H:%M")
            chave = f"selecionar_historico_{pasta.name}"

            with st.container(border=True):
                col_sel, col_info, col_download = st.columns([0.05, 0.50, 0.45])

                with col_sel:
                    marcado = st.checkbox(
                        "Selecionar histórico",
                        key=chave,
                        label_visibility="collapsed",
                    )

                with col_info:
                    st.markdown(
                        f"""
                        <div style="padding-top:6px; line-height:1.25;">
                            <div style="font-weight:650; font-size:15px;">
                                {pasta.name}
                            </div>
                            <div style="opacity:0.68; font-size:12px; padding-top:3px;">
                                {data_txt}
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                with col_download:
                    col_prev, col_excel, col_pdf = st.columns([1.1, 1, 1])

                    if item["excel"] and item["excel"].exists():
                        excel_bytes = ler_bytes(item["excel"])
                        if excel_bytes:
                            col_excel.download_button(
                                "Excel",
                                data=excel_bytes,
                                file_name=item["excel"].name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"hist_excel_{pasta.name}",
                                use_container_width=True,
                            )

                    if item["pdf"] and item["pdf"].exists():
                        pdf_bytes = ler_bytes(item["pdf"])
                        if pdf_bytes:
                            col_pdf.download_button(
                                "PDF",
                                data=pdf_bytes,
                                file_name=item["pdf"].name,
                                mime="application/pdf",
                                key=f"hist_pdf_{pasta.name}",
                                use_container_width=True,
                            )

                    with col_prev:
                        abrir_previa = st.toggle(
                            "Prévia",
                            key=f"toggle_previa_{pasta.name}",
                        )

                if abrir_previa:
                    _renderizar_previa_historico_excel(
                        item.get("excel"),
                        chave=f"previa_{pasta.name}",
                    )

                if marcado:
                    selecionados.append(pasta)

        _renderizar_paginacao("pagina_convertidos", pagina_atual, total_paginas)

        st.caption(f"Selecionados nesta página: {len(selecionados)}")

        if st.button(
            "🗑️ Excluir convertidos selecionados",
            type="secondary",
            use_container_width=True,
            disabled=len(selecionados) == 0,
            key="excluir_convertidos_selecionados",
        ):
            qtd = excluir_itens_historico(selecionados)

            excel_atual = Path(st.session_state.get("excel", "")) if st.session_state.get("excel") else None
            pdf_atual = Path(st.session_state.get("pdf", "")) if st.session_state.get("pdf") else None

            for pasta in selecionados:
                pasta = Path(pasta)
                st.session_state.pop(f"selecionar_historico_{pasta.name}", None)
                st.session_state.pop(f"toggle_previa_{pasta.name}", None)

                if (excel_atual and pasta in excel_atual.parents) or (pdf_atual and pasta in pdf_atual.parents):
                    for chave_estado in ["df_final", "info", "excel", "pdf", "nome_saida", "percentual", "resumo_processamento"]:
                        st.session_state.pop(chave_estado, None)

            st.success(f"{qtd} item(ns) excluído(s) do histórico.")
            st.rerun()

def extrair_placa_falha(linha):
    """Extrai a placa de uma linha/registro que falhou no processamento."""
    texto = str(linha or "")
    match = re.search(r"[A-Z]{3}[0-9][A-Z0-9][0-9]{2}", texto.upper())
    return match.group(0) if match else "Placa não identificada"


def resumir_falhas_processamento(falhas):
    """Agrupa falhas por placa para facilitar conferência na tela."""
    resumo = {}

    for falha in falhas or []:
        placa = extrair_placa_falha(falha)
        resumo.setdefault(placa, [])
        resumo[placa].append(str(falha))

    return resumo


def renderizar_falhas_processamento(falhas):
    """Mostra na interface apenas as placas que deram falha, com detalhes opcionais."""
    if not falhas:
        return

    resumo = resumir_falhas_processamento(falhas)
    total_placas = len(resumo)
    total_linhas = len(falhas)

    with st.expander(f"⚠️ Ver placas com falha ({total_placas})", expanded=False):
        st.caption(
            f"{total_linhas} linha(s) não foram processadas automaticamente. "
            "Confira as placas abaixo e, se necessário, abra os detalhes."
        )

        for placa, linhas in resumo.items():
            st.markdown(f"**{placa}**")
            with st.container(border=True):
                for linha in linhas[:3]:
                    st.code(linha, language="text")
                if len(linhas) > 3:
                    st.caption(f"+ {len(linhas) - 3} ocorrência(s) adicional(is) dessa placa.")




def montar_zip_resultados_individuais(resultados, tipo="todos"):
    """
    Monta um ZIP com os arquivos individuais, respeitando os nomes editados na tela.

    tipo:
    - "excel": inclui somente arquivos .xlsx
    - "pdf": inclui somente arquivos .pdf
    - "todos": inclui Excel + PDF
    """
    buffer = io.BytesIO()
    nomes_usados = {}
    tipo = str(tipo or "todos").lower().strip()

    if tipo not in {"excel", "pdf", "todos"}:
        tipo = "todos"

    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zipf:
        for idx_resultado, resultado in enumerate(resultados or []):
            df_final = resultado.get("df_final")

            if not isinstance(df_final, pd.DataFrame) or df_final.empty:
                continue

            arquivo_original = resultado.get("arquivo_original", "arquivo")
            nome_padrao = resultado.get("nome_saida", Path(arquivo_original).stem)
            chave_nome_download = f"nome_download_resultado_{idx_resultado}_{nome_seguro(arquivo_original)}"
            nome_download = nome_seguro(st.session_state.get(chave_nome_download, nome_padrao))

            qtd_nome = nomes_usados.get(nome_download, 0) + 1
            nomes_usados[nome_download] = qtd_nome

            if qtd_nome > 1:
                nome_download_zip = f"{nome_download}_{qtd_nome}"
            else:
                nome_download_zip = nome_download

            caminho_excel = Path(resultado["excel"]) if resultado.get("excel") else None
            caminho_pdf = Path(resultado["pdf"]) if resultado.get("pdf") else None

            excel_bytes = ler_bytes(caminho_excel) if caminho_excel else None
            pdf_bytes = ler_bytes(caminho_pdf) if caminho_pdf else None

            if tipo in {"excel", "todos"} and excel_bytes:
                zipf.writestr(f"{nome_download_zip}.xlsx", excel_bytes)

            if tipo in {"pdf", "todos"} and pdf_bytes:
                zipf.writestr(f"{nome_download_zip}.pdf", pdf_bytes)

    buffer.seek(0)
    return buffer.getvalue()

def main():
    st.set_page_config(
        page_title="R3R Intermediações",
        page_icon="assets/icon.png",
        layout="wide",
    )

    st.markdown(
        """
        <style>
            section[data-testid="stSidebar"] {
                display: none !important;
            }

            div[data-testid="collapsedControl"] {
                display: none !important;
            }

            .block-container {
                padding-top: 0.8rem !important;
                padding-left: 2rem !important;
                padding-right: 2rem !important;
                max-width: 100% !important;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )

    limpar_historico_antigo(dias=10)

    st.markdown("<div style='height:35px'></div>", unsafe_allow_html=True)

    col1, col2 = st.columns([0.56, 0.44])

    with col1:
        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
        banner_path = Path("assets/text_icon.png")

        if banner_path.exists():
            st.image(str(banner_path), width=620)
        else:
            st.error("Imagem text_icon.png não encontrada em assets/")

    with col2:
        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
        logo_path = obter_logo_streamlit_path()

        if logo_path:
            st.image(str(logo_path), width=380)

    st.markdown("<div style='height:-10px'></div>", unsafe_allow_html=True)

    with st.container(border=True):
        st.markdown("### Ajuste de margem")
        percentual = st.number_input(
            "Acréscimo sobre o preço (%)",
            min_value=0.0,
            max_value=100.0,
            value=4.0,
            step=0.5,
            format="%.2f",
            help="Percentual aplicado sobre o preço base para recalcular o preço final.",
        )
        st.caption("Esse percentual será usado para recalcular preço final, distância FIPE e margem final.")

    with st.container(border=True):
        st.markdown("### 📤 Arquivos de entrada")
        st.caption(
            "Envie uma ou mais tabelas em PDF, XLSX ou XLS. "
            "Cada arquivo será convertido individualmente, com prévia, nome próprio e downloads separados."
        )

        uploaded_files = st.file_uploader(
            "Selecione ou arraste o(s) arquivo(s) aqui",
            type=["pdf", "xlsx", "xls"],
            accept_multiple_files=True,
            help="Envie um ou mais PDFs/planilhas. O sistema processará cada lista separadamente.",
            label_visibility="collapsed",
        )

        if uploaded_files:
            st.success(f"{len(uploaded_files)} arquivo(s) carregado(s).")
            with st.expander("Ver arquivos selecionados", expanded=False):
                for arquivo in uploaded_files:
                    st.caption(f"• {arquivo.name}")

    if not uploaded_files:
        st.info("Envie um arquivo para iniciar.")
        renderizar_historico()
        return

    with st.container(border=True):
        st.markdown("### ⬇️ Nome individual de cada lista")
        st.caption(
            "Defina o nome final de cada conversão antes de gerar os arquivos. "
            "Cada item terá seu próprio Excel e PDF."
        )

        nomes_saida_por_arquivo = []

        for idx_arquivo, arquivo in enumerate(uploaded_files):
            nome_base = Path(arquivo.name).stem
            chave_nome = f"nome_saida_individual_{idx_arquivo}_{nome_seguro(arquivo.name)}"

            nome_digitado = st.text_input(
                f"Nome final — {arquivo.name}",
                value=f"{nome_base}_margem_atualizada",
                key=chave_nome,
                placeholder="Digite o nome final desta lista",
            )

            nomes_saida_por_arquivo.append(nome_seguro(nome_digitado))

    if st.button("Gerar listas individuais com margem", type="primary", use_container_width=True):
        resultados_individuais = []
        resumo_processamento = []
        nomes_usados = {}

        with st.spinner("Processando arquivo(s) individualmente..."):
            for idx_arquivo, arquivo in enumerate(uploaded_files):
                nome_saida = nomes_saida_por_arquivo[idx_arquivo]

                # Evita sobrescrever arquivos caso dois nomes finais sejam iguais no mesmo lote.
                qtd_nome = nomes_usados.get(nome_saida, 0) + 1
                nomes_usados[nome_saida] = qtd_nome
                if qtd_nome > 1:
                    nome_saida = f"{nome_saida}_{qtd_nome}"

                try:
                    caminho_upload = salvar_upload(arquivo)
                    df_final, info = processar_arquivo(
                        caminho_upload,
                        percentual=percentual,
                    )

                    falhas = info.get("falhas", []) or []

                    if df_final.empty:
                        resumo_processamento.append({
                            "ARQUIVO": arquivo.name,
                            "NOME FINAL": nome_saida,
                            "MODO": str(info.get("modo", "-")).upper(),
                            "REGISTROS": 0,
                            "FALHAS": len(falhas),
                            "STATUS": "SEM REGISTROS",
                        })

                        resultados_individuais.append({
                            "arquivo_original": arquivo.name,
                            "nome_saida": nome_saida,
                            "df_final": pd.DataFrame(),
                            "info": info,
                            "excel": None,
                            "pdf": None,
                            "erro": None,
                        })
                        continue

                    caminho_excel = salvar_excel(df_final, nome_arquivo=f"{nome_saida}.xlsx")
                    caminho_pdf = salvar_pdf(df_final, nome_arquivo=f"{nome_saida}.pdf")
                    hist_excel, hist_pdf = salvar_historico(caminho_excel, caminho_pdf, nome_saida)

                    resumo_processamento.append({
                        "ARQUIVO": arquivo.name,
                        "NOME FINAL": nome_saida,
                        "MODO": str(info.get("modo", "-")).upper(),
                        "REGISTROS": len(df_final),
                        "FALHAS": len(falhas),
                        "STATUS": "OK",
                    })

                    resultados_individuais.append({
                        "arquivo_original": arquivo.name,
                        "nome_saida": nome_saida,
                        "df_final": df_final,
                        "info": info,
                        "excel": str(hist_excel),
                        "pdf": str(hist_pdf),
                        "erro": None,
                    })

                except Exception as erro_arquivo:
                    resumo_processamento.append({
                        "ARQUIVO": arquivo.name,
                        "NOME FINAL": nome_saida,
                        "MODO": "ERRO",
                        "REGISTROS": 0,
                        "FALHAS": 1,
                        "STATUS": str(erro_arquivo),
                    })

                    resultados_individuais.append({
                        "arquivo_original": arquivo.name,
                        "nome_saida": nome_saida,
                        "df_final": pd.DataFrame(),
                        "info": {"modo": "erro", "falhas": [str(erro_arquivo)]},
                        "excel": None,
                        "pdf": None,
                        "erro": str(erro_arquivo),
                    })

        st.session_state["resultados_individuais"] = resultados_individuais
        st.session_state["resumo_processamento_individual"] = pd.DataFrame(resumo_processamento)
        st.session_state["percentual"] = percentual

        # Limpa estados antigos do fluxo consolidado, caso existam.
        for chave_estado in ["df_final", "info", "excel", "pdf", "nome_saida", "resumo_processamento"]:
            st.session_state.pop(chave_estado, None)

        qtd_ok = sum(
            1 for r in resultados_individuais
            if isinstance(r.get("df_final"), pd.DataFrame) and not r["df_final"].empty
        )

        st.success(f"{qtd_ok} lista(s) individual(is) processada(s) com sucesso.")

    resultados = st.session_state.get("resultados_individuais", [])

    if resultados:
        resultados_validos = [
            r for r in resultados
            if isinstance(r.get("df_final"), pd.DataFrame) and not r["df_final"].empty
        ]

        if resultados_validos:
            with st.container(border=True):
                st.markdown("### 📦 Download em lote")
                st.caption(
                    "Baixe os arquivos individuais em lote. "
                    "Você pode baixar somente Excel ou somente PDF. "
                    "Os nomes usados serão os nomes editados em cada lista."
                )

                timestamp_zip = datetime.now().strftime('%Y%m%d_%H%M%S')
                col_zip_excel, col_zip_pdf = st.columns(2)

                with col_zip_excel:
                    excel_icon = Path("assets/excel_icon.png")
                    if excel_icon.exists():
                        st.markdown(
                            f"""
                            <div style="text-align:center; margin-bottom:8px;">
                                <img src="data:image/png;base64,{base64.b64encode(excel_icon.read_bytes()).decode()}" width="42">
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )

                    zip_excel_bytes = montar_zip_resultados_individuais(resultados_validos, tipo="excel")
                    st.download_button(
                        "⬇️ Baixar todos Excel (.zip)",
                        data=zip_excel_bytes,
                        file_name=f"listas_excel_{timestamp_zip}.zip",
                        mime="application/zip",
                        use_container_width=True,
                        key="download_zip_excel_individuais_topo",
                    )

                with col_zip_pdf:
                    pdf_icon = Path("assets/pdf_icon.png")
                    if pdf_icon.exists():
                        st.markdown(
                            f"""
                            <div style="text-align:center; margin-bottom:8px;">
                                <img src="data:image/png;base64,{base64.b64encode(pdf_icon.read_bytes()).decode()}" width="42">
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )

                    zip_pdf_bytes = montar_zip_resultados_individuais(resultados_validos, tipo="pdf")
                    st.download_button(
                        "⬇️ Baixar todos PDF (.zip)",
                        data=zip_pdf_bytes,
                        file_name=f"listas_pdf_{timestamp_zip}.zip",
                        mime="application/zip",
                        use_container_width=True,
                        key="download_zip_pdf_individuais_topo",
                    )

        resumo_processamento = st.session_state.get("resumo_processamento_individual")

        if isinstance(resumo_processamento, pd.DataFrame) and not resumo_processamento.empty:
            with st.expander("Resumo por arquivo processado", expanded=True):
                st.dataframe(resumo_processamento, use_container_width=True, height=220)

        st.markdown("## Listas convertidas individualmente")

        def formatar_preview_percentual(x):
            try:
                valor = round(float(x) * 100, 2)

                if valor.is_integer():
                    return f"{int(valor)}%"

                texto = f"{valor:.2f}".rstrip("0").rstrip(".")
                return f"{texto}%".replace(".", ",")

            except Exception:
                return x

        for idx_resultado, resultado in enumerate(resultados):
            arquivo_original = resultado.get("arquivo_original", "arquivo")
            nome_padrao_resultado = resultado.get("nome_saida", Path(arquivo_original).stem)
            df_final = resultado.get("df_final", pd.DataFrame())
            info = resultado.get("info", {}) or {}
            erro = resultado.get("erro")

            titulo = f"📄 {arquivo_original}"

            if isinstance(df_final, pd.DataFrame) and not df_final.empty:
                titulo += f" — {len(df_final)} registro(s)"
            elif erro:
                titulo += " — erro"
            else:
                titulo += " — sem registros"

            with st.expander(titulo, expanded=True):
                if erro:
                    st.error(f"Erro ao processar este arquivo: {erro}")
                    renderizar_falhas_processamento(info.get("falhas", []))
                    continue

                if not isinstance(df_final, pd.DataFrame) or df_final.empty:
                    st.warning("Nenhum registro válido foi extraído desta lista.")
                    renderizar_falhas_processamento(info.get("falhas", []))
                    continue

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Registros", f"{len(df_final):,}".replace(",", "."))
                c2.metric("Modo", str(info.get("modo", "-")).upper())
                c3.metric("Falhas", len(info.get("falhas", [])))
                c4.metric("Acréscimo", f"{st.session_state.get('percentual', percentual):.2f}%".replace(".", ","))

                chave_nome_download = f"nome_download_resultado_{idx_resultado}_{nome_seguro(arquivo_original)}"
                nome_download = st.text_input(
                    "Renomear antes do download",
                    value=nome_padrao_resultado,
                    key=chave_nome_download,
                    help="Esse nome será usado nos botões de download desta lista.",
                )
                nome_download = nome_seguro(nome_download)

                st.markdown("#### Prévia")
                preview_df = df_final.copy()

                if "MARGEM FINAL" in preview_df.columns:
                    preview_df["MARGEM FINAL"] = preview_df["MARGEM FINAL"].apply(
                        formatar_preview_percentual
                    )

                st.dataframe(preview_df, use_container_width=True, height=360)

                st.markdown("#### Downloads desta lista")
                d0, d1, d2, d3 = st.columns([0.22, 0.28, 0.28, 0.22])

                caminho_excel = Path(resultado["excel"]) if resultado.get("excel") else None
                caminho_pdf = Path(resultado["pdf"]) if resultado.get("pdf") else None

                excel_bytes = ler_bytes(caminho_excel) if caminho_excel else None
                pdf_bytes = ler_bytes(caminho_pdf) if caminho_pdf else None

                if excel_bytes:
                    with d1:
                        excel_icon = Path("assets/excel_icon.png")
                        if excel_icon.exists():
                            st.markdown(
                                f"""
                                <div style="text-align:center; margin-bottom:8px;">
                                    <img src="data:image/png;base64,{base64.b64encode(excel_icon.read_bytes()).decode()}" width="42">
                                </div>
                                """,
                                unsafe_allow_html=True,
                            )

                        if st.download_button(
                            "Baixar Excel",
                            data=excel_bytes,
                            file_name=f"{nome_download}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_excel_individual_{idx_resultado}",
                            use_container_width=True,
                        ):
                            st.toast("Download do Excel iniciado.", icon="✅")
                else:
                    d1.warning("Excel não encontrado.")

                if pdf_bytes:
                    with d2:
                        pdf_icon = Path("assets/pdf_icon.png")
                        if pdf_icon.exists():
                            st.markdown(
                                f"""
                                <div style="text-align:center; margin-bottom:8px;">
                                    <img src="data:image/png;base64,{base64.b64encode(pdf_icon.read_bytes()).decode()}" width="42">
                                </div>
                                """,
                                unsafe_allow_html=True,
                            )

                        if st.download_button(
                            "Baixar PDF",
                            data=pdf_bytes,
                            file_name=f"{nome_download}.pdf",
                            mime="application/pdf",
                            key=f"download_pdf_individual_{idx_resultado}",
                            use_container_width=True,
                        ):
                            st.toast("Download do PDF iniciado.", icon="✅")
                else:
                    d2.warning("PDF não encontrado.")

                renderizar_falhas_processamento(info.get("falhas", []))


    renderizar_historico_uploads()
    renderizar_historico()


if __name__ == "__main__":
    main()
